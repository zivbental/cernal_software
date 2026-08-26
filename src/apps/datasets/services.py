"""Dataset upload and validation.

Validation here is **shallow and synchronous**: readable file, expected columns, sane
row count, parseable numerics. Deep scientific validation belongs to the engine and
happens at run time (docs/architecture.md §5). Duplicating scientific truth on the
Platform side is exactly what design map 04 warns against.
"""

import csv
import io
import logging
from collections.abc import Iterator
from pathlib import Path

from django.conf import settings
from django.core.files.uploadedfile import SimpleUploadedFile
from django.db import transaction

from apps.common.checksums import sha256_upload
from apps.datasets.models import Dataset, ValidationStatus

logger = logging.getLogger(__name__)

#: Canonical column name -> the spellings researchers actually export.
#: DESeq2, edgeR, limma and Excel all disagree, and none of them are wrong.
COLUMN_ALIASES: dict[str, tuple[str, ...]] = {
    "gene_id": ("gene_id", "gene", "geneid", "gene_name", "genename", "id", "symbol", "target_id"),
    "log2fc": (
        "log2fc",
        "log2foldchange",
        "log2_fold_change",
        "logfc",
        "fold_change",
        "foldchange",
        "fc",
    ),
    "base_expression": ("base_expression", "basemean", "base_mean", "control", "control_mean"),
    "target_expression": ("target_expression", "target", "target_mean", "treatment"),
    "padj": ("padj", "p_adj", "adj_pval", "adj_p_val", "fdr", "qvalue", "q_value"),
    "pvalue": ("pvalue", "p_value", "pval", "p"),
}

#: Every dataset must identify its features.
REQUIRED_COLUMNS = frozenset({"gene_id"})

#: At least one of these must be present, and must parse as a number.
EXPRESSION_COLUMNS = ("log2fc", "base_expression", "target_expression")

#: Columns validated as numeric when present.
NUMERIC_COLUMNS = frozenset({*EXPRESSION_COLUMNS, "padj", "pvalue"})

#: What we will attempt to parse.
SUPPORTED_SUFFIXES = (".csv", ".tsv", ".txt", ".xlsx")

MAX_ROWS = 200_000
SAMPLE_ROWS = 5_000


#: Bundled datasets a researcher can try the product with, so evaluating CERNAL does
#: not require having your own differential-expression results to hand.
EXAMPLES: dict[str, dict[str, str]] = {
    "ecoli-oxidative-stress": {
        "filename": "ecoli_oxidative_stress.csv",
        "label": "E. coli — lactose metabolism to oxidative stress",
        "description": (
            "50 genes from a differential-expression analysis comparing standard growth "
            "with oxidative stress. Twenty stress-response genes are up-regulated, "
            "fifteen metabolic and motility genes down, and fifteen housekeeping genes "
            "are unchanged."
        ),
    },
}

EXAMPLES_DIR = Path(__file__).resolve().parent / "examples"


class DatasetValidationError(Exception):
    """The upload could not be accepted at all (as opposed to failing validation)."""


@transaction.atomic
def create_example_dataset(*, project, user, key: str = "ecoli-oxidative-stress"):
    """Attach a bundled example dataset to a project.

    Goes through exactly the same checksum and validation path as an upload, so an
    example run is indistinguishable from a real one downstream.
    """
    try:
        example = EXAMPLES[key]
    except KeyError:
        known = ", ".join(sorted(EXAMPLES))
        raise DatasetValidationError(
            f"Unknown example dataset '{key}'. Available: {known}."
        ) from None

    source = EXAMPLES_DIR / example["filename"]
    if not source.is_file():
        raise DatasetValidationError("That example dataset is missing from this install.")

    return create_dataset(
        project=project,
        uploaded_file=SimpleUploadedFile(
            example["filename"], source.read_bytes(), content_type="text/csv"
        ),
        user=user,
        name=example["filename"],
    )


@transaction.atomic
def create_dataset(*, project, uploaded_file, user, name: str | None = None) -> Dataset:
    """Store an upload, checksum it, and validate it — all before returning.

    The dataset is immutable once written, so everything that can be known about it is
    established here.
    """
    size = uploaded_file.size
    limit = settings.MAX_DATASET_MB * 1024 * 1024
    if size > limit:
        raise DatasetValidationError(
            f"The file is larger than the {settings.MAX_DATASET_MB} MB limit."
        )
    if size == 0:
        raise DatasetValidationError("The file is empty.")

    checksum = sha256_upload(uploaded_file)
    report = validate_expression_file(uploaded_file)
    uploaded_file.seek(0)

    dataset = Dataset(
        project=project,
        name=name or uploaded_file.name,
        checksum_sha256=checksum,
        size_bytes=size,
        schema_version="1",
        validation_status=(
            ValidationStatus.VALID if not report["errors"] else ValidationStatus.INVALID
        ),
        validation_report=report,
        uploaded_by=user,
    )
    dataset.file.save(dataset.name, uploaded_file, save=False)
    dataset.save()

    logger.info(
        "Dataset %s uploaded to project %s: %s (%d rows)",
        dataset.id,
        project.id,
        dataset.validation_status,
        report.get("rows", 0),
    )
    return dataset


def validate_expression_file(uploaded_file) -> dict:
    """Inspect a CSV upload. Returns a report; never raises for bad *content*.

    Errors mean the file cannot be analysed. Warnings mean it can, but something looks
    off and the researcher should know.
    """
    report: dict = {"rows": 0, "columns": [], "errors": [], "warnings": []}

    try:
        raw = uploaded_file.read()
    finally:
        uploaded_file.seek(0)

    filename = getattr(uploaded_file, "name", "") or ""

    try:
        columns, rows_iter = _read_table(raw, filename)
    except DatasetValidationError as exc:
        report["errors"].append(str(exc))
        return report

    report["columns"] = columns
    if not columns:
        report["errors"].append("The file has no header row.")
        return report

    canonical = _canonical_columns(columns)
    report["detected_columns"] = canonical
    present = set(canonical.values())
    reader = rows_iter

    missing = sorted(REQUIRED_COLUMNS - present)
    if missing:
        report["errors"].append(f"Missing required column(s): {', '.join(missing)}.")

    available_expression = [name for name in EXPRESSION_COLUMNS if name in present]
    if not available_expression:
        report["errors"].append(
            "No fold-change or expression column found. Recognised names include: "
            "log2FoldChange, log2fc, fold_change, FC, baseMean."
        )

    numeric_present = sorted(NUMERIC_COLUMNS & present)
    bad_numbers: list[str] = []
    seen_ids: set[str] = set()
    duplicates = 0
    rows = 0

    for index, row in enumerate(reader, start=2):
        rows += 1
        if rows > MAX_ROWS:
            report["errors"].append(f"The file has more than {MAX_ROWS:,} rows.")
            break

        if rows <= SAMPLE_ROWS:
            row = {canonical.get(k, k): v for k, v in row.items()}
            gene_id = (row.get("gene_id") or "").strip()
            if gene_id:
                if gene_id in seen_ids:
                    duplicates += 1
                seen_ids.add(gene_id)

            for column in numeric_present:
                value = (row.get(column) or "").strip()
                if value in ("", "NA", "NaN", "null"):
                    continue
                try:
                    float(value)
                except ValueError:
                    if len(bad_numbers) < 5:
                        bad_numbers.append(f"row {index}, column '{column}': {value!r}")

    report["rows"] = rows

    if rows == 0:
        report["errors"].append("The file has a header but no data rows.")
    if bad_numbers:
        report["errors"].append(
            "Non-numeric values in numeric columns — " + "; ".join(bad_numbers) + "."
        )
    if duplicates:
        report["warnings"].append(
            f"{duplicates} duplicate gene_id value(s) in the first {SAMPLE_ROWS:,} rows."
        )
    if rows > SAMPLE_ROWS:
        report["warnings"].append(f"Only the first {SAMPLE_ROWS:,} rows were checked in detail.")

    return report


def _canonical_columns(columns: list[str]) -> dict[str, str]:
    """Map each header in the file to the canonical name we understand.

    Comparison ignores case, spaces, dots and underscores, because
    ``log2 Fold Change``, ``log2FoldChange`` and ``log2_fold_change`` are the same column.
    """

    def normalize(name: str) -> str:
        return "".join(ch for ch in name.lower() if ch.isalnum())

    lookup = {
        normalize(alias): canonical
        for canonical, aliases in COLUMN_ALIASES.items()
        for alias in aliases
    }
    return {name: lookup.get(normalize(name), name.strip().lower()) for name in columns}


def _read_table(raw: bytes, filename: str) -> tuple[list[str], Iterator[dict]]:
    """Return ``(headers, row dicts)`` for a CSV, TSV or XLSX upload.

    Format is chosen by extension, then confirmed by content — a ``.csv`` that is really
    a spreadsheet is a common and confusing mistake.
    """
    suffix = Path(filename).suffix.lower()

    if raw[:2] == b"PK" or suffix == ".xlsx":
        if suffix not in (".xlsx", ""):
            raise DatasetValidationError(
                f"This looks like an Excel workbook but is named '{filename}'. "
                "Rename it to .xlsx, or export it as CSV."
            )
        return _read_xlsx(raw)

    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        raise DatasetValidationError(
            "The file is not valid UTF-8 text. Export it as CSV, TSV or XLSX and retry."
        ) from None

    delimiter = "\t" if suffix in (".tsv", ".txt") else _sniff_delimiter(text)
    reader = csv.DictReader(io.StringIO(text), delimiter=delimiter)
    headers = [name.strip() for name in (reader.fieldnames or [])]
    return headers, reader


def _sniff_delimiter(text: str) -> str:
    """Guess the separator from the header line. Falls back to a comma."""
    header = text.split("\n", 1)[0]
    try:
        return csv.Sniffer().sniff(header, delimiters=",\t;").delimiter
    except csv.Error:
        return "\t" if header.count("\t") > header.count(",") else ","


def _read_xlsx(raw: bytes) -> tuple[list[str], Iterator[dict]]:
    """Read the first worksheet. Streamed, so a large workbook does not sit in memory."""
    from openpyxl import load_workbook

    try:
        workbook = load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    except Exception:
        raise DatasetValidationError("The Excel workbook could not be opened.") from None

    sheet = workbook.worksheets[0]
    rows = sheet.iter_rows(values_only=True)

    try:
        header_row = next(rows)
    except StopIteration:
        return [], iter(())

    headers = [str(cell).strip() if cell is not None else "" for cell in header_row]

    def as_dicts() -> Iterator[dict]:
        for row in rows:
            if all(cell is None for cell in row):
                continue
            yield {
                headers[i]: ("" if value is None else str(value))
                for i, value in enumerate(row)
                if i < len(headers)
            }
        workbook.close()

    return headers, as_dicts()


def delete_dataset(dataset) -> None:
    """Remove a dataset that no run has used.

    A dataset referenced by a run is PROTECTed at the database level; this turns that
    into a clear message rather than an IntegrityError.
    """
    if dataset.runs.exists():
        raise DatasetValidationError(
            "This dataset has been analysed and cannot be deleted. "
            "Results must not outlive the input that produced them."
        )
    dataset.delete()
